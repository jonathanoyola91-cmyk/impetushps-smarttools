from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from cotizador.models import CurvaOperacion, EquipoBase


class Command(BaseCommand):
    help = "Importa equipos base y curvas de operación desde un archivo Excel."

    def add_arguments(self, parser):
        parser.add_argument(
            "excel_path",
            type=str,
            help="Ruta del archivo Excel a importar",
        )
        parser.add_argument(
            "--replace",
            action="store_true",
            help="Borra los registros existentes antes de importar",
        )

    def handle(self, *args, **options):
        excel_path = Path(options["excel_path"])
        replace = options["replace"]

        if not excel_path.exists():
            raise CommandError(f"No existe el archivo: {excel_path}")

        try:
            wb = load_workbook(excel_path, data_only=True)
        except Exception as exc:
            raise CommandError(f"No se pudo abrir el Excel: {exc}") from exc

        required_sheets = ["equipos_base", "curvas_operacion"]
        missing = [name for name in required_sheets if name not in wb.sheetnames]
        if missing:
            raise CommandError(
                f"Faltan hojas requeridas en el Excel: {', '.join(missing)}"
            )

        with transaction.atomic():
            if replace:
                self.stdout.write(self.style.WARNING("Borrando datos anteriores..."))
                CurvaOperacion.objects.all().delete()
                EquipoBase.objects.all().delete()

            equipos_creados, equipos_actualizados = self._importar_equipos_base(
                wb["equipos_base"]
            )
            curvas_creadas, curvas_actualizadas = self._importar_curvas_operacion(
                wb["curvas_operacion"]
            )

        self.stdout.write(self.style.SUCCESS("Importación completada correctamente."))
        self.stdout.write(
            self.style.SUCCESS(
                f"Equipos base -> creados: {equipos_creados}, actualizados: {equipos_actualizados}"
            )
        )
        self.stdout.write(
            self.style.SUCCESS(
                f"Curvas operación -> creadas: {curvas_creadas}, actualizadas: {curvas_actualizadas}"
            )
        )

    def _importar_equipos_base(self, ws):
        headers = self._get_headers(ws)

        required_headers = [
            "id_equipo",
            "nombre_equipo",
            "bomba",
            "camara_empuje",
            "carga_camara_max",
            "longitud_total_mm",
            "potencia_hp_min",
            "potencia_hp_max",
            "caudal_min_bpd",
            "caudal_max_bpd",
            "ps_min_psi",
            "ps_max_psi",
            "pd_min_psi",
            "pd_max_psi",
            "precio_base",
            "tiempo_base_dias",
        ]
        self._validate_headers("equipos_base", headers, required_headers)

        created = 0
        updated = 0

        for row_number, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if self._row_is_empty(row):
                continue

            data = dict(zip(headers, row))

            id_equipo = self._clean_str(data.get("id_equipo"))
            if not id_equipo:
                raise CommandError(
                    f"Hoja equipos_base, fila {row_number}: id_equipo es obligatorio."
                )

            defaults = {
                "nombre_equipo": self._clean_str(data.get("nombre_equipo")),
                "bomba": self._clean_str(data.get("bomba")),
                "camara_empuje": self._clean_str(data.get("camara_empuje")),
                "carga_camara_max": self._to_decimal(data.get("carga_camara_max")),
                "longitud_total_mm": self._to_decimal(data.get("longitud_total_mm")),
                "potencia_hp_min": self._to_decimal_required(
                    data.get("potencia_hp_min"),
                    "equipos_base",
                    row_number,
                    "potencia_hp_min",
                ),
                "potencia_hp_max": self._to_decimal_required(
                    data.get("potencia_hp_max"),
                    "equipos_base",
                    row_number,
                    "potencia_hp_max",
                ),
                "caudal_min_bpd": self._to_decimal_required(
                    data.get("caudal_min_bpd"),
                    "equipos_base",
                    row_number,
                    "caudal_min_bpd",
                ),
                "caudal_max_bpd": self._to_decimal_required(
                    data.get("caudal_max_bpd"),
                    "equipos_base",
                    row_number,
                    "caudal_max_bpd",
                ),
                "ps_min_psi": self._to_decimal_required(
                    data.get("ps_min_psi"),
                    "equipos_base",
                    row_number,
                    "ps_min_psi",
                ),
                "ps_max_psi": self._to_decimal_required(
                    data.get("ps_max_psi"),
                    "equipos_base",
                    row_number,
                    "ps_max_psi",
                ),
                "pd_min_psi": self._to_decimal_required(
                    data.get("pd_min_psi"),
                    "equipos_base",
                    row_number,
                    "pd_min_psi",
                ),
                "pd_max_psi": self._to_decimal_required(
                    data.get("pd_max_psi"),
                    "equipos_base",
                    row_number,
                    "pd_max_psi",
                ),
                "precio_base": self._to_decimal(data.get("precio_base")) or Decimal("0"),
                "tiempo_base_dias": self._to_int(data.get("tiempo_base_dias")) or 0,
                "activo": True,
            }

            obj, was_created = EquipoBase.objects.update_or_create(
                id_equipo=id_equipo,
                defaults=defaults,
            )

            if was_created:
                created += 1
            else:
                updated += 1

        return created, updated

    def _importar_curvas_operacion(self, ws):
        headers = self._get_headers(ws)

        required_headers = [
            "id_punto",
            "id_equipo",
            "caudal",
            "dp_requerida",
            "potencia_hp",
            "carga_camara",
            "eficiencia_pct",
            "tipo_punto",
        ]
        self._validate_headers("curvas_operacion", headers, required_headers)

        allowed_tipo_punto = {"NOMINAL", "BAJA_EFICIENCIA", "ALTO_CAUDAL"}

        created = 0
        updated = 0

        for row_number, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if self._row_is_empty(row):
                continue

            data = dict(zip(headers, row))

            id_punto = self._clean_str(data.get("id_punto"))
            id_equipo = self._clean_str(data.get("id_equipo"))

            if not id_punto:
                raise CommandError(
                    f"Hoja curvas_operacion, fila {row_number}: id_punto es obligatorio."
                )
            if not id_equipo:
                raise CommandError(
                    f"Hoja curvas_operacion, fila {row_number}: id_equipo es obligatorio."
                )

            try:
                equipo = EquipoBase.objects.get(id_equipo=id_equipo)
            except EquipoBase.DoesNotExist as exc:
                raise CommandError(
                    f"Hoja curvas_operacion, fila {row_number}: no existe el equipo '{id_equipo}' en equipos_base."
                ) from exc

            tipo_punto = self._clean_str(data.get("tipo_punto")).upper()
            if tipo_punto not in allowed_tipo_punto:
                raise CommandError(
                    f"Hoja curvas_operacion, fila {row_number}: tipo_punto '{tipo_punto}' no válido."
                )

            defaults = {
                "equipo": equipo,
                "caudal": self._to_decimal_required(
                    data.get("caudal"),
                    "curvas_operacion",
                    row_number,
                    "caudal",
                ),
                "dp_requerida": self._to_decimal_required(
                    data.get("dp_requerida"),
                    "curvas_operacion",
                    row_number,
                    "dp_requerida",
                ),
                "potencia_hp": self._to_decimal_required(
                    data.get("potencia_hp"),
                    "curvas_operacion",
                    row_number,
                    "potencia_hp",
                ),
                "carga_camara": self._to_decimal_required(
                    data.get("carga_camara"),
                    "curvas_operacion",
                    row_number,
                    "carga_camara",
                ),
                "eficiencia_pct": self._to_decimal_required(
                    data.get("eficiencia_pct"),
                    "curvas_operacion",
                    row_number,
                    "eficiencia_pct",
                ),
                "tipo_punto": tipo_punto,
            }

            obj, was_created = CurvaOperacion.objects.update_or_create(
                id_punto=id_punto,
                defaults=defaults,
            )

            if was_created:
                created += 1
            else:
                updated += 1

        return created, updated

    def _get_headers(self, ws):
        headers = []
        for cell in ws[1]:
            headers.append(self._clean_str(cell.value))
        return headers

    def _validate_headers(self, sheet_name, headers, required_headers):
        missing = [h for h in required_headers if h not in headers]
        if missing:
            raise CommandError(
                f"En la hoja '{sheet_name}' faltan columnas requeridas: {', '.join(missing)}"
            )

    def _clean_str(self, value):
        if value is None:
            return ""
        return str(value).strip()

    def _to_decimal(self, value):
        if value is None or value == "":
            return None
        try:
            text = str(value).strip().replace(",", ".")
            return Decimal(text)
        except (InvalidOperation, ValueError):
            return None

    def _to_decimal_required(self, value, sheet_name, row_number, field_name):
        decimal_value = self._to_decimal(value)
        if decimal_value is None:
            raise CommandError(
                f"Hoja {sheet_name}, fila {row_number}: valor inválido en '{field_name}' -> {value}"
            )
        return decimal_value

    def _to_int(self, value):
        if value is None or value == "":
            return None
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None

    def _row_is_empty(self, row):
        return all(cell is None or str(cell).strip() == "" for cell in row)