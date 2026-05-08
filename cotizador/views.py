from decimal import Decimal, InvalidOperation
import json

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from django.db.models import Q
from django.http import FileResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone

from .email_utils import (
    enviar_cotizacion_por_correo,
    notificar_nueva_cotizacion_bomba,
    notificar_nueva_reparacion_camara,
    notificar_precio_reparacion_camara,
    notificar_uso_sospechoso_vsd,
)
from .forms import ImportadorDiagnosticoVariadorForm, ImportadorReparacionCamaraForm
from .models import (
    ConsultaDiagnosticoVariadorLog,
    DiagnosticoVariador,
    ReparacionCamaraTarifa,
    SolicitudCotizacion,
    SolicitudReparacionCamara,
)
from .pdf_utils import generar_pdf_cotizacion
from .services import seleccionar_mejor_punto, calcular_dp

from django.contrib import messages
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.shortcuts import render, redirect
from django.conf import settings

from .forms import SolicitudCuentaForm

from django.contrib.auth import logout
from django.shortcuts import redirect

def logout_view(request):
    logout(request)
    return redirect("home")

def crear_cuenta(request):
    if request.method == "POST":
        form = SolicitudCuentaForm(request.POST)

        if form.is_valid():
            nombre = form.cleaned_data["nombre"]
            empresa = form.cleaned_data["empresa"]
            email = form.cleaned_data["email"]
            username = form.cleaned_data["username"]
            password = form.cleaned_data["password"]

            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=nombre,
                is_active=False,
            )

            asunto = "Nueva solicitud de acceso - IMPETUS Smart Tools"

            mensaje = f"""
Nueva solicitud de acceso registrada.

Nombre: {nombre}
Empresa: {empresa}
Correo: {email}
Usuario: {username}

El usuario quedó pendiente de autorización.
Debe activarse manualmente desde el panel administrador.
"""

            send_mail(
                asunto,
                mensaje,
                settings.DEFAULT_FROM_EMAIL,
                ["director.comercial@impetushps.co"],
                fail_silently=True,
            )

            messages.success(
                request,
                "Solicitud enviada correctamente. Su cuenta será revisada por IMPETUS."
            )

            return redirect("login")

    else:
        form = SolicitudCuentaForm()

    return render(
        request,
        "cotizador/crear_cuenta.html",
        {"form": form}
    )
VSD_MONITOR_WINDOW_MINUTES = 60
VSD_MONITOR_IP_THRESHOLD = 25
VSD_MONITOR_SESSION_THRESHOLD = 18


def _str_value(value):
    return str(value).strip() if value is not None else ""


def _decimal_value(value, default="0"):
    if value is None or str(value).strip() == "":
        return Decimal(default)
    try:
        return Decimal(str(value).replace("$", "").replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return Decimal(default)


def _bool_value(value, default=True):
    if value is None or str(value).strip() == "":
        return default
    return str(value).strip().lower() in ["si", "sí", "true", "1", "activo", "yes", "x"]


def _client_ip(request):
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded_for:
        return x_forwarded_for.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR")


def _registrar_consulta_vsd(request, marca, codigo, encontrado):
    if not request.session.session_key:
        request.session.save()

    log = ConsultaDiagnosticoVariadorLog.objects.create(
        marca=marca,
        codigo=codigo,
        encontrado=encontrado,
        ip_address=_client_ip(request),
        user_agent=request.META.get("HTTP_USER_AGENT", "")[:2000],
        session_key=request.session.session_key or "",
        referer=request.META.get("HTTP_REFERER", "")[:2000],
        path=request.path[:255],
    )

    desde = timezone.now() - timezone.timedelta(minutes=VSD_MONITOR_WINDOW_MINUTES)
    por_ip = ConsultaDiagnosticoVariadorLog.objects.filter(
        ip_address=log.ip_address,
        created_at__gte=desde,
    ).count() if log.ip_address else 0
    por_sesion = ConsultaDiagnosticoVariadorLog.objects.filter(
        session_key=log.session_key,
        created_at__gte=desde,
    ).count() if log.session_key else 0

    if por_ip >= VSD_MONITOR_IP_THRESHOLD or por_sesion >= VSD_MONITOR_SESSION_THRESHOLD:
        log.es_sospechoso = True
        try:
            notificar_uso_sospechoso_vsd(log, por_ip, por_sesion, VSD_MONITOR_WINDOW_MINUTES)
            log.notificacion_enviada = True
        except Exception:
            log.notificacion_enviada = False
        log.save(update_fields=["es_sospechoso", "notificacion_enviada"])

    return log


def home_view(request):
    return render(request, "cotizador/home.html")


def reparacion_camara_view(request):
    resultado = None
    error = None
    solicitud_id = None

    tarifas = ReparacionCamaraTarifa.objects.filter(activo=True).order_by("marca", "modelo")
    marcas = list(tarifas.values_list("marca", flat=True).distinct())
    modelos_por_marca = {}

    for marca in marcas:
        modelos_por_marca[marca] = list(
            tarifas.filter(marca=marca).values_list("modelo", flat=True).distinct()
        )

    if request.method == "POST":
        if not request.user.is_authenticated:
            messages.warning(
                request,
                "Debe crear una cuenta o iniciar sesión para calcular la reparación."
            )
            return redirect("crear_cuenta")

        try:
            if request.user.is_authenticated:
                empresa = request.POST.get("empresa", "").strip() or request.user.first_name or request.user.username
                contacto = request.POST.get("contacto", "").strip() or request.user.get_full_name() or request.user.username
                correo = request.POST.get("correo", "").strip() or request.user.email
                telefono = request.POST.get("telefono", "").strip()
            else:
                empresa = request.POST.get("empresa", "").strip()
                contacto = request.POST.get("contacto", "").strip()
                correo = request.POST.get("correo", "").strip()
                telefono = request.POST.get("telefono", "").strip()
            nombre_proyecto = request.POST.get("nombre_proyecto", "").strip()
            marca = request.POST.get("marca", "").strip()
            modelo = request.POST.get("modelo", "").strip()
            serial = request.POST.get("serial", "").strip()
            tipo_reparacion = request.POST.get("tipo_reparacion", "").strip()
            observaciones_cliente = request.POST.get("observaciones_cliente", "").strip()

            if not empresa or not contacto or not correo:
                error = "No fue posible identificar empresa, contacto o correo del usuario."
            else:
                tarifa = ReparacionCamaraTarifa.objects.filter(
                    activo=True,
                    marca=marca,
                    modelo=modelo,
                    tipo_reparacion=tipo_reparacion,
                ).first()

                if not tarifa:
                    error = "No se encontró configuración para esa selección."
                else:
                    solicitud = SolicitudReparacionCamara.objects.create(
                        empresa=empresa,
                        contacto=contacto,
                        correo=correo,
                        telefono=telefono,
                        nombre_proyecto=nombre_proyecto,
                        marca=marca,
                        modelo=modelo,
                        serial=serial,
                        tipo_reparacion=tipo_reparacion,
                        observaciones_cliente=observaciones_cliente,
                        observacion_tecnica=tarifa.observacion,
                        tiempo_estimado_texto=tarifa.tiempo_estimado_texto,
                        valor_estimado=tarifa.valor_estimado,
                    )
                    resultado = solicitud
                    solicitud_id = solicitud.id
                    try:
                        notificar_nueva_reparacion_camara(solicitud)
                    except Exception as exc:
                        messages.warning(request, f"Solicitud creada, pero no se pudo enviar notificación: {exc}")
        except Exception as exc:
            error = f"Error: {exc}"

    return render(request, "cotizador/reparacion_camara.html", {
        "resultado": resultado,
        "error": error,
        "solicitud_id": solicitud_id,
        "marcas": marcas,
        "modelos_por_marca_json": json.dumps(modelos_por_marca),
    })


@login_required
def solicitar_precio_reparacion(request, solicitud_id):
    solicitud = get_object_or_404(SolicitudReparacionCamara, id=solicitud_id)
    solicitud.solicito_precio = True
    solicitud.fecha_solicitud_precio = timezone.now()
    solicitud.save()

    try:
        notificar_precio_reparacion_camara(solicitud)
    except Exception as exc:
        messages.warning(request, f"Precio solicitado, pero no se pudo enviar notificación: {exc}")

    mensaje = (
        f"El cliente {solicitud.empresa} solicitó precio para la reparación "
        f"de cámara {solicitud.marca} {solicitud.modelo} - {solicitud.tipo_reparacion}."
    )
    return render(request, "cotizador/precio_solicitado.html", {"solicitud": solicitud, "mensaje": mensaje})


@login_required
def descargar_pdf(request, solicitud_id):
    solicitud = get_object_or_404(SolicitudCotizacion, id=solicitud_id)
    ruta = f"cotizacion_{solicitud.id}.pdf"
    generar_pdf_cotizacion(ruta, solicitud)
    return FileResponse(open(ruta, "rb"), as_attachment=True)


@login_required
def enviar_cotizacion_email(request, solicitud_id):
    solicitud = get_object_or_404(SolicitudCotizacion, id=solicitud_id)
    enviar_cotizacion_por_correo(
        solicitud,
        correo_copia=getattr(settings, "COTIZADOR_EMAIL_COPIA", "director.comercial@impetushps.co"),
    )

    numero_cotizacion = f"CP-{solicitud.id:04d}"
    es_modo_prueba = settings.EMAIL_BACKEND == "django.core.mail.backends.console.EmailBackend"
    if es_modo_prueba:
        mensaje = f"La cotización presupuestal {numero_cotizacion} fue procesada en modo prueba."
    else:
        mensaje = f"La cotización presupuestal {numero_cotizacion} fue enviada correctamente al correo {solicitud.correo}."

    return render(request, "cotizador/email_enviado.html", {
        "solicitud": solicitud,
        "mensaje": mensaje,
        "numero_cotizacion": numero_cotizacion,
        "es_modo_prueba": es_modo_prueba,
    })


def cotizador_view(request):
    resultado = None
    error = None
    solicitud_id = None

    if request.method == "POST":
        if not request.user.is_authenticated:
            messages.warning(
                request,
                "Debe crear una cuenta o iniciar sesión para calcular la cotización."
            )
            return redirect("crear_cuenta")

        try:
            empresa = request.POST.get("empresa", "").strip() or request.user.first_name or request.user.username
            contacto = request.POST.get("contacto", "").strip() or request.user.get_full_name() or request.user.username
            correo = request.POST.get("correo", "").strip() or request.user.email or "director.comercial@impetushps.co"
            telefono = request.POST.get("telefono", "").strip()
            
            nombre_proyecto = request.POST.get("nombre_proyecto", "").strip()
            marca = request.POST.get("marca", "").strip()
            modelo = request.POST.get("modelo", "").strip()
            serial = request.POST.get("serial", "").strip()
            tipo_reparacion = request.POST.get("tipo_reparacion", "").strip()
            observaciones_cliente = request.POST.get("observaciones_cliente", "").strip()
           

            if not empresa or not contacto or not correo:
                error = "No fue posible identificar empresa, contacto o correo del usuario."
            else:
                tarifa = ReparacionCamaraTarifa.objects.filter(
                    activo=True,
                    marca=marca,
                    modelo=modelo,
                    tipo_reparacion=tipo_reparacion,
                ).first()

                if not tarifa:
                    error = "No se encontró configuración para esa selección."
                else:
                    solicitud = SolicitudReparacionCamara.objects.create(
                        empresa=empresa,
                        contacto=contacto,
                        correo=correo,
                        telefono=telefono,
                        nombre_proyecto=nombre_proyecto,
                        marca=marca,
                        modelo=modelo,
                        serial=serial,
                        tipo_reparacion=tipo_reparacion,
                        observaciones_cliente=observaciones_cliente,
                        observacion_tecnica=tarifa.observacion,
                        tiempo_estimado_texto=tarifa.tiempo_estimado_texto,
                        valor_estimado=tarifa.valor_estimado,
                    )

                    resultado = solicitud
                    solicitud_id = solicitud.id

                    try:
                        notificar_nueva_reparacion_camara(solicitud)
                    except Exception as exc:
                        messages.warning(request, f"Solicitud creada, pero no se pudo enviar notificación: {exc}")

        except Exception as exc:
            error = f"Error: {exc}"

    return render(request, "cotizador/formulario.html", {
        "resultado": resultado,
        "error": error,
        "solicitud_id": solicitud_id,
    })


@login_required
def historial_cotizaciones(request):
    q = request.GET.get("q", "").strip()
    solicitudes = SolicitudCotizacion.objects.all().order_by("-created_at")
    if q:
        solicitudes = solicitudes.filter(
            Q(empresa__icontains=q) | Q(contacto__icontains=q) | Q(correo__icontains=q) | Q(nombre_proyecto__icontains=q)
        ).order_by("-created_at")
    return render(request, "cotizador/historial.html", {"solicitudes": solicitudes, "q": q})


def diagnostico_variador_view(request):
    resultado = None
    error = None

    marcas = DiagnosticoVariador.objects.filter(activo=True).values_list("marca", flat=True).distinct().order_by("marca")

    if request.method == "POST":
        marca = request.POST.get("marca", "").strip()
        codigo = request.POST.get("codigo", "").strip()
        website = request.POST.get("website", "").strip()  # honeypot anti-bot simple

        if website:
            error = "No se pudo procesar la consulta."
            _registrar_consulta_vsd(request, marca, codigo, False)
        else:
            resultado = DiagnosticoVariador.objects.filter(activo=True, marca__iexact=marca, codigo__iexact=codigo).first()
            _registrar_consulta_vsd(request, marca, codigo, bool(resultado))
            if not resultado:
                error = "No se encontró información para ese código."

    return render(request, "cotizador/diagnostico_variador.html", {"resultado": resultado, "error": error, "marcas": marcas})


def _procesar_importacion_diagnostico(archivo):
    wb = openpyxl.load_workbook(archivo, data_only=True)
    ws = wb.active
    encabezados_requeridos = ["Marca", "Código", "Tipo", "Nombre de Falla", "Causa Probable", "Acción Recomendada", "Categoría", "Activo"]
    encabezados_archivo = [_str_value(cell.value) for cell in ws[1]]
    faltantes = [col for col in encabezados_requeridos if col not in encabezados_archivo]
    if faltantes:
        raise ValueError(f"Faltan columnas obligatorias: {', '.join(faltantes)}")
    indices = {col: encabezados_archivo.index(col) for col in encabezados_requeridos}
    creados = actualizados = filas_vacias = 0
    errores = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        marca = _str_value(row[indices["Marca"]])
        codigo = _str_value(row[indices["Código"]])
        tipo = _str_value(row[indices["Tipo"]])
        nombre_falla = _str_value(row[indices["Nombre de Falla"]])
        causa_probable = _str_value(row[indices["Causa Probable"]])
        accion_recomendada = _str_value(row[indices["Acción Recomendada"]])
        categoria = _str_value(row[indices["Categoría"]])
        activo = _bool_value(row[indices["Activo"]], default=True)
        if not marca and not codigo and not tipo and not nombre_falla:
            filas_vacias += 1
            continue
        if not marca or not codigo or not tipo or not nombre_falla:
            errores.append(f"Fila {idx}: faltan campos obligatorios.")
            continue
        _, creado = DiagnosticoVariador.objects.update_or_create(
            marca=marca,
            codigo=codigo,
            defaults={"tipo": tipo, "nombre_falla": nombre_falla, "causa_probable": causa_probable, "accion_recomendada": accion_recomendada, "categoria": categoria, "activo": activo},
        )
        creados += 1 if creado else 0
        actualizados += 0 if creado else 1
    return {"creados": creados, "actualizados": actualizados, "filas_vacias": filas_vacias, "errores": errores}


@login_required
def importar_diagnostico_variador_view(request):
    resumen = None
    error = None
    mensaje_ok = None
    form = ImportadorDiagnosticoVariadorForm(request.POST or None, request.FILES or None)

    if request.method == "POST":
        if form.is_valid():
            archivo = form.cleaned_data["archivo"]
            if not archivo.name.lower().endswith(".xlsx"):
                error = "Debe cargar un archivo Excel .xlsx"
            else:
                try:
                    resumen = _procesar_importacion_diagnostico(archivo)
                    mensaje_ok = f"Importación completada. Creados: {resumen['creados']} | Actualizados: {resumen['actualizados']} | Filas vacías: {resumen['filas_vacias']}"
                except Exception as exc:
                    error = f"Error procesando archivo: {exc}"
        else:
            error = "Debe seleccionar un archivo válido."

    return render(request, "cotizador/importar_diagnostico_variador.html", {"form": form, "resumen": resumen, "error": error, "mensaje_ok": mensaje_ok})


def _procesar_importacion_reparacion_camara(archivo):
    wb = openpyxl.load_workbook(archivo, data_only=True)
    ws = wb.active
    encabezados_requeridos = ["Marca", "Modelo", "Tipo Reparacion", "Valor Estimado", "Tiempo Estimado", "Observacion", "Activo"]
    encabezados_archivo = [_str_value(cell.value) for cell in ws[1]]
    faltantes = [col for col in encabezados_requeridos if col not in encabezados_archivo]
    if faltantes:
        raise ValueError(f"Faltan columnas obligatorias: {', '.join(faltantes)}")
    indices = {col: encabezados_archivo.index(col) for col in encabezados_requeridos}
    creados = actualizados = filas_vacias = 0
    errores = []
    tipos_validos = {"MENOR", "MAYOR", "UPGRADE"}

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        marca = _str_value(row[indices["Marca"]])
        modelo = _str_value(row[indices["Modelo"]])
        tipo_reparacion = _str_value(row[indices["Tipo Reparacion"]]).upper()
        valor_estimado = _decimal_value(row[indices["Valor Estimado"]])
        tiempo_estimado = _str_value(row[indices["Tiempo Estimado"]])
        observacion = _str_value(row[indices["Observacion"]])
        activo = _bool_value(row[indices["Activo"]], default=True)

        if not marca and not modelo and not tipo_reparacion:
            filas_vacias += 1
            continue
        if not marca or not modelo or not tipo_reparacion:
            errores.append(f"Fila {idx}: Marca, Modelo y Tipo Reparacion son obligatorios.")
            continue
        if tipo_reparacion not in tipos_validos:
            errores.append(f"Fila {idx}: Tipo Reparacion debe ser MENOR, MAYOR o UPGRADE.")
            continue

        _, creado = ReparacionCamaraTarifa.objects.update_or_create(
            marca=marca,
            modelo=modelo,
            tipo_reparacion=tipo_reparacion,
            defaults={"valor_estimado": valor_estimado, "tiempo_estimado_texto": tiempo_estimado, "observacion": observacion, "activo": activo},
        )
        creados += 1 if creado else 0
        actualizados += 0 if creado else 1
    return {"creados": creados, "actualizados": actualizados, "filas_vacias": filas_vacias, "errores": errores}


@login_required
def importar_reparacion_camara_view(request):
    resumen = None
    error = None
    mensaje_ok = None
    form = ImportadorReparacionCamaraForm(request.POST or None, request.FILES or None)

    if request.method == "POST":
        if form.is_valid():
            archivo = form.cleaned_data["archivo"]
            if not archivo.name.lower().endswith(".xlsx"):
                error = "Debe cargar un archivo Excel .xlsx"
            else:
                try:
                    resumen = _procesar_importacion_reparacion_camara(archivo)
                    mensaje_ok = f"Importación completada. Creados: {resumen['creados']} | Actualizados: {resumen['actualizados']} | Filas vacías: {resumen['filas_vacias']}"
                except Exception as exc:
                    error = f"Error procesando archivo: {exc}"
        else:
            error = "Debe seleccionar un archivo válido."

    return render(request, "cotizador/importar_reparacion_camara.html", {"form": form, "resumen": resumen, "error": error, "mensaje_ok": mensaje_ok})



class CustomLoginView(LoginView):
    template_name = "cotizador/login.html"
    redirect_authenticated_user = True

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["mensaje_acceso"] = (
            "Para usar los módulos de Bombas y Reparación de Cámaras debe iniciar sesión. "
            "El diagnóstico VSD permanece público para soporte técnico rápido."
        )
        return context
